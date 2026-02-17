"""Calc-category fixture file generators (xlsx, ods, csv, tsv)."""

import csv
from pathlib import Path


# ---------------------------------------------------------------------------
# Shared test data — "월별 매출 보고서" scenario (FR-017)
# ---------------------------------------------------------------------------

CALC_HEADERS = ["월", "지역", "매출액", "비용", "순이익"]

CALC_REGIONS = ["서울", "부산", "대구"]

CALC_MONTHS = [
    "1월",
    "2월",
    "3월",
    "4월",
    "5월",
    "6월",
    "7월",
    "8월",
    "9월",
    "10월",
    "11월",
    "12월",
]

# Generate realistic-looking sales data (12 months × 3 regions = 36 rows)
_BASE_SALES = [
    [12000, 8500, 6200],  # Jan
    [13500, 9200, 7100],  # Feb
    [15000, 10000, 7800],  # Mar
    [14200, 9500, 7400],  # Apr
    [16000, 11000, 8500],  # May
    [17500, 12000, 9200],  # Jun
    [18000, 12500, 9800],  # Jul
    [16500, 11500, 8800],  # Aug
    [15500, 10500, 8200],  # Sep
    [14800, 9800, 7600],  # Oct
    [13200, 8800, 6800],  # Nov
    [19000, 13000, 10000],  # Dec
]

_COST_RATIO = 0.65  # Cost is ~65% of sales


def _build_rows():
    """Build the 36-row data set: (month, region, sales, cost, profit)."""
    rows = []
    for m_idx, month in enumerate(CALC_MONTHS):
        for r_idx, region in enumerate(CALC_REGIONS):
            sales = _BASE_SALES[m_idx][r_idx]
            cost = int(sales * _COST_RATIO)
            profit = sales - cost
            rows.append([month, region, sales, cost, profit])
    return rows


CALC_DATA_ROWS = _build_rows()


# ---------------------------------------------------------------------------
# T016: create_xlsx
# ---------------------------------------------------------------------------


def create_xlsx(path: Path) -> Path:
    """Create a .xlsx file with headers, data rows, and formulas.

    Uses the "월별 매출 보고서" (Monthly Sales Report) scenario (FR-017).
    Contains 36 data rows (12 months × 3 regions) plus SUM/AVERAGE formulas.

    Args:
        path: File path to create.

    Returns:
        The created file Path.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "매출 보고서"

    # Headers (bold)
    for col_idx, header in enumerate(CALC_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, row_data in enumerate(CALC_DATA_ROWS, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Formula rows (after data, row 38 = row 2 + 36 data rows)
    last_data_row = 1 + len(CALC_DATA_ROWS)  # 37
    formula_row = last_data_row + 1  # 38

    ws.cell(row=formula_row, column=1, value="합계")
    ws.cell(row=formula_row, column=3, value=f"=SUM(C2:C{last_data_row})")
    ws.cell(row=formula_row, column=4, value=f"=SUM(D2:D{last_data_row})")
    ws.cell(row=formula_row, column=5, value=f"=SUM(E2:E{last_data_row})")

    avg_row = formula_row + 1
    ws.cell(row=avg_row, column=1, value="평균")
    ws.cell(row=avg_row, column=3, value=f"=AVERAGE(C2:C{last_data_row})")
    ws.cell(row=avg_row, column=4, value=f"=AVERAGE(D2:D{last_data_row})")
    ws.cell(row=avg_row, column=5, value=f"=AVERAGE(E2:E{last_data_row})")

    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# T017: create_ods
# ---------------------------------------------------------------------------


def create_ods(path: Path) -> Path:
    """Create a .ods file with headers and data rows.

    Requires odfpy. Caller should handle ImportError via pytest.importorskip.

    Args:
        path: File path to create.

    Returns:
        The created file Path.
    """
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = OpenDocumentSpreadsheet()
    table = Table(name="매출 보고서")

    # Header row
    header_row = TableRow()
    for header in CALC_HEADERS:
        cell = TableCell()
        cell.addElement(P(text=header))
        header_row.addElement(cell)
    table.addElement(header_row)

    # Data rows
    for row_data in CALC_DATA_ROWS:
        row = TableRow()
        for value in row_data:
            cell = TableCell()
            cell.addElement(P(text=str(value)))
            row.addElement(cell)
        table.addElement(row)

    doc.spreadsheet.addElement(table)
    doc.save(str(path))
    return path


# ---------------------------------------------------------------------------
# T018: create_csv
# ---------------------------------------------------------------------------


def create_csv_file(path: Path) -> Path:
    """Create a .csv file with comma-separated sales data.

    Args:
        path: File path to create.

    Returns:
        The created file Path.
    """
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(CALC_HEADERS)
        writer.writerows(CALC_DATA_ROWS)
    return path


# ---------------------------------------------------------------------------
# T019: create_tsv
# ---------------------------------------------------------------------------


def create_tsv(path: Path) -> Path:
    """Create a .tsv file with tab-separated sales data.

    Args:
        path: File path to create.

    Returns:
        The created file Path.
    """
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerow(CALC_HEADERS)
        writer.writerows(CALC_DATA_ROWS)
    return path


# ---------------------------------------------------------------------------
# Edge case helpers (Phase 7)
# ---------------------------------------------------------------------------


def create_empty_xlsx(path: Path) -> Path:
    """Create an empty .xlsx with one empty sheet (FR-015)."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.save(str(path))
    return path


def create_large_xlsx(path: Path) -> Path:
    """Create a .xlsx with 1000+ rows (FR-015)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Large Data"

    # Headers
    headers = ["ID", "Name", "Value", "Category", "Score"]
    ws.append(headers)

    # 1200 data rows
    for i in range(1, 1201):
        ws.append(
            [
                i,
                f"Item_{i:04d}",
                i * 1.5,
                f"Cat_{(i % 10) + 1}",
                (i * 17) % 100,
            ]
        )

    wb.save(str(path))
    return path
